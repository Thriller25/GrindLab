import logging
import uuid
from io import BytesIO
from typing import Any, Optional

from fastapi import APIRouter, Depends, HTTPException, Response
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session

from app.db import get_db
from app.schemas.calc_run import CalcRunCommentUpdate, CalcRunCreate, CalcRunRead
from app.schemas.grind_mvp import (
    GrindMvpInput,
    GrindMvpResult,
    GrindMvpRunDetail,
    GrindMvpRunResponse,
    GrindMvpRunSummary,
)
from app.services.calc_service import (
    CalculationError,
    get_flowsheet_version_or_404,
    run_grind_mvp_calculation,
    run_flowsheet_calculation,
    run_flowsheet_calculation_by_scenario,
)
from app.routers.auth import get_current_user, get_current_user_optional
from app import models

router = APIRouter(prefix="/api/calc", tags=["calc"])
logger = logging.getLogger(__name__)


@router.post("/flowsheet-run", response_model=CalcRunRead)
def calc_flowsheet(payload: CalcRunCreate, db: Session = Depends(get_db)) -> CalcRunRead:
    """
    Run comminution flowsheet calculation and persist CalcRun metadata.
    """
    try:
        get_flowsheet_version_or_404(db, payload.flowsheet_version_id)
        return run_flowsheet_calculation(db=db, payload=payload)
    except CalculationError as exc:
        raise HTTPException(status_code=400, detail=str(exc))
    except HTTPException:
        raise
    except Exception:
        logger.exception("Internal calculation error")
        raise HTTPException(status_code=500, detail="Internal calculation error")


def _extract_model_version(run: models.CalcRun) -> Optional[str]:
    if isinstance(run.result_json, dict):
        mv = run.result_json.get("model_version")
        if mv:
            return mv
    if isinstance(run.input_json, dict):
        mv = run.input_json.get("model_version")
        if mv:
            return mv
    return None


def _is_grind_mvp_run(run: models.CalcRun) -> bool:
    return _extract_model_version(run) == "grind_mvp_v1"


def _safe_to_str(value: Any) -> Optional[str]:
    if value is None:
        return None
    return str(value)


def _build_grind_mvp_summary(run: models.CalcRun) -> GrindMvpRunSummary:
    input_json = run.input_json if isinstance(run.input_json, dict) else {}
    result_json = run.result_json if isinstance(run.result_json, dict) else {}
    kpi = result_json.get("kpi") if isinstance(result_json, dict) else None
    throughput = kpi.get("throughput_tph") if isinstance(kpi, dict) else None
    product_p80 = kpi.get("product_p80_mm") if isinstance(kpi, dict) else None
    spec_energy = kpi.get("specific_energy_kwh_per_t") if isinstance(kpi, dict) else None

    return GrindMvpRunSummary(
        id=run.id,
        created_at=run.created_at,
        model_version=_extract_model_version(run) or "grind_mvp_v1",
        plant_id=_safe_to_str(input_json.get("plant_id")),
        flowsheet_version_id=_safe_to_str(input_json.get("flowsheet_version_id")),
        scenario_name=run.scenario_name or input_json.get("scenario_name"),
        comment=run.comment,
        throughput_tph=throughput,
        product_p80_mm=product_p80,
        specific_energy_kwhpt=spec_energy,
    )


def _build_grind_mvp_detail(calc_run: models.CalcRun) -> GrindMvpRunDetail:
    if calc_run is None or not _is_grind_mvp_run(calc_run):
        raise HTTPException(status_code=404, detail="Calc run not found")

    if not isinstance(calc_run.result_json, dict):
        raise HTTPException(status_code=404, detail="Calc run not found")

    result_model = GrindMvpResult.model_validate(calc_run.result_json)
    input_json = calc_run.input_json if isinstance(calc_run.input_json, dict) else None
    if not input_json:
        raise HTTPException(status_code=404, detail="Calc run not found")
    input_model = GrindMvpInput.model_validate(input_json)

    return GrindMvpRunDetail(
        id=calc_run.id,
        created_at=calc_run.created_at,
        model_version=_extract_model_version(calc_run) or result_model.model_version,
        plant_id=_safe_to_str(input_json.get("plant_id")),
        flowsheet_version_id=_safe_to_str(input_json.get("flowsheet_version_id")),
        scenario_name=calc_run.scenario_name or input_json.get("scenario_name"),
        comment=calc_run.comment,
        input=input_model,
        result=result_model,
    )


@router.get("/grind-mvp-runs", response_model=list[GrindMvpRunSummary])
def list_grind_mvp_runs(
    limit: int = 20,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user_optional),
):
    runs = db.query(models.CalcRun).order_by(models.CalcRun.created_at.desc()).all()
    grind_runs = [run for run in runs if _is_grind_mvp_run(run)]
    summaries = [_build_grind_mvp_summary(run) for run in grind_runs[:limit]]
    return summaries


@router.get("/grind-mvp-runs/{run_id}", response_model=GrindMvpRunDetail)
def get_grind_mvp_run(
    run_id: uuid.UUID,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user_optional),
):
    calc_run = db.query(models.CalcRun).filter(models.CalcRun.id == run_id).first()
    return _build_grind_mvp_detail(calc_run)


@router.put("/grind-mvp-runs/{run_id}/comment", response_model=GrindMvpRunDetail)
def update_grind_mvp_run_comment(
    run_id: uuid.UUID,
    payload: CalcRunCommentUpdate,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user_optional),
):
    calc_run = db.query(models.CalcRun).filter(models.CalcRun.id == run_id).first()
    if calc_run is None or not _is_grind_mvp_run(calc_run):
        raise HTTPException(status_code=404, detail="Calc run not found")

    calc_run.comment = payload.comment
    db.commit()
    db.refresh(calc_run)
    return _build_grind_mvp_detail(calc_run)


@router.get("/grind-mvp-runs/{run_id}/report.xlsx")
def get_grind_mvp_run_report(
    run_id: uuid.UUID,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user_optional),
):
    """
    Excel-отчёт по расчёту grind_mvp_v1.
    """
    calc_run = db.query(models.CalcRun).filter(models.CalcRun.id == run_id).first()
    if calc_run is None or not _is_grind_mvp_run(calc_run):
        raise HTTPException(status_code=404, detail="Calc run not found")
    if not isinstance(calc_run.result_json, dict) or not isinstance(calc_run.input_json, dict):
        raise HTTPException(status_code=404, detail="Calc run not found")

    input_model = GrindMvpInput.model_validate(calc_run.input_json)
    result_model = GrindMvpResult.model_validate(calc_run.result_json)

    wb = Workbook()
    ws_summary = wb.active
    ws_summary.title = "Сводка"

    row = 1
    ws_summary[f"A{row}"] = "Расчёт №"
    ws_summary[f"B{row}"] = str(calc_run.id)
    row += 1
    ws_summary[f"A{row}"] = "Модель"
    ws_summary[f"B{row}"] = input_model.model_version
    row += 1
    ws_summary[f"A{row}"] = "Фабрика ID"
    ws_summary[f"B{row}"] = str(input_model.plant_id)
    row += 1
    ws_summary[f"A{row}"] = "Версия схемы ID"
    ws_summary[f"B{row}"] = str(input_model.flowsheet_version_id)
    row += 1
    ws_summary[f"A{row}"] = "Сценарий"
    ws_summary[f"B{row}"] = input_model.scenario_name or ""
    row += 1
    ws_summary[f"A{row}"] = "Создан"
    ws_summary[f"B{row}"] = calc_run.created_at.isoformat() if calc_run.created_at else ""
    row += 2

    ws_summary[f"A{row}"] = "Ключевые показатели"
    row += 1
    ws_summary[f"A{row}"] = "Производительность, т/ч"
    ws_summary[f"B{row}"] = result_model.kpi.throughput_tph
    row += 1
    ws_summary[f"A{row}"] = "P80 продукта, мм"
    ws_summary[f"B{row}"] = result_model.kpi.product_p80_mm
    row += 1
    ws_summary[f"A{row}"] = "Удельная энергия, кВт·ч/т"
    ws_summary[f"B{row}"] = result_model.kpi.specific_energy_kwh_per_t
    row += 1
    ws_summary[f"A{row}"] = "Циркуляционная нагрузка, %"
    ws_summary[f"B{row}"] = result_model.kpi.circulating_load_percent
    row += 1
    ws_summary[f"A{row}"] = "Использование мощности, %"
    ws_summary[f"B{row}"] = result_model.kpi.mill_utilization_percent
    row += 2

    if result_model.baseline_comparison:
        bc = result_model.baseline_comparison
        ws_summary[f"A{row}"] = "Сравнение с базовым сценарием"
        row += 1
        ws_summary[f"A{row}"] = "Базовый расчёт №"
        ws_summary[f"B{row}"] = str(bc.baseline_run_id)
        row += 1
        ws_summary[f"A{row}"] = "Δ Производительность, т/ч"
        ws_summary[f"B{row}"] = bc.throughput_delta_tph
        row += 1
        ws_summary[f"A{row}"] = "Δ P80 продукта, мм"
        ws_summary[f"B{row}"] = bc.product_p80_delta_mm
        row += 1
        ws_summary[f"A{row}"] = "Δ Удельная энергия, кВт·ч/т"
        ws_summary[f"B{row}"] = bc.specific_energy_delta_kwhpt
        row += 2

    ws_sd = wb.create_sheet(title="Грансостав")
    row = 1
    ws_sd[f"A{row}"] = "Гранулометрический состав питания"
    row += 1
    ws_sd[f"A{row}"] = "Размер, мм"
    ws_sd[f"B{row}"] = "Сумм., %"
    row += 1
    for point in result_model.size_distribution.feed:
        ws_sd[f"A{row}"] = point.size_mm
        ws_sd[f"B{row}"] = point.cum_percent
        row += 1

    row += 1
    ws_sd[f"A{row}"] = "Гранулометрический состав продукта"
    row += 1
    ws_sd[f"A{row}"] = "Размер, мм"
    ws_sd[f"B{row}"] = "Сумм., %"
    row += 1
    for point in result_model.size_distribution.product:
        ws_sd[f"A{row}"] = point.size_mm
        ws_sd[f"B{row}"] = point.cum_percent
        row += 1

    for ws in (ws_summary, ws_sd):
        for col in range(1, 3):
            ws.column_dimensions[get_column_letter(col)].width = 28

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"grind_run_{calc_run.id}.xlsx"
    headers = {"Content-Disposition": f'attachment; filename=\"{filename}\"'}

    return Response(
        content=buffer.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers,
    )

@router.post("/flowsheet-run/by-scenario/{scenario_id}", response_model=CalcRunRead)
def calc_flowsheet_by_scenario(scenario_id: uuid.UUID, db: Session = Depends(get_db)) -> CalcRunRead:
    """
    Run calculation using default input stored on CalcScenario.
    """
    try:
        return run_flowsheet_calculation_by_scenario(db=db, scenario_id=scenario_id)
    except CalculationError as exc:
        raise HTTPException(status_code=400, detail=str(exc))
    except HTTPException:
        raise
    except Exception:
        logger.exception("Internal calculation error")
        raise HTTPException(status_code=500, detail="Internal calculation error")


@router.post("/grind-mvp-runs", response_model=GrindMvpRunResponse)
def calc_grind_mvp_run(
    payload: GrindMvpInput,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user_optional),
) -> GrindMvpRunResponse:
    """
    Grind MVP calculation entrypoint using model_version='grind_mvp_v1'.
    Requires authentication.
    """
    try:
        result = run_grind_mvp_calculation(db=db, payload=payload)
        return GrindMvpRunResponse.model_validate(
            {"calc_run_id": result["calc_run_id"], "result": result["result"]}
        )
    except CalculationError as exc:
        raise HTTPException(status_code=400, detail=str(exc))
    except HTTPException:
        raise
    except Exception:
        logger.exception("Internal calculation error")
        raise HTTPException(status_code=500, detail="Internal calculation error")
